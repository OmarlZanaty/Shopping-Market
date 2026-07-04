"""
Synchronous product import from XLSX/CSV, keyed on barcode.

Semantics:
- barcode is required on every row (the upsert key).
- Existing barcode  -> update; blank cells keep the current value (partial update).
- Unknown barcode   -> create; requires name_ar and original_price.
- `categories` column: comma-separated category names (matches name_en or
  name_ar, case-insensitive, within the store). Unmatched names produce a
  warning, not an error — the row still imports.
- dry_run computes the same result without writing anything.

Result shape:
    {
        'total_rows': int, 'created': int, 'updated': int,
        'errors':   [{'row': int, 'barcode': str, 'reason': str}],
        'warnings': [{'row': int, 'barcode': str, 'reason': str}],
    }
"""
from decimal import Decimal, InvalidOperation

from django.db import transaction

from .models import Product, Category

MAX_ROWS = 50000

# Model fields writable via import; bulk_update rewrites all of them for every
# touched row using the instance's in-memory value (unchanged fields keep the
# value loaded from the DB, so listing them all here is safe).
UPDATE_FIELDS = [
    'name_ar', 'name_en', 'description_ar', 'description_en',
    'original_price', 'discount_price', 'cost_price',
    'quantity_in_stock', 'low_stock_threshold', 'sell_unit',
    'is_weight_based', 'is_available', 'image_url_s3', 'main_image',
]

# header -> model field (identity unless noted)
COLUMNS = [
    'barcode', 'name_ar', 'name_en', 'description_ar', 'description_en',
    'categories', 'original_price', 'discount_price', 'cost_price',
    'quantity_in_stock', 'low_stock_threshold', 'sell_unit',
    'is_weight_based', 'is_available', 'image_url',
]

REQUIRED_FOR_CREATE = ('name_ar', 'original_price')

_TRUE = {'1', 'true', 'yes', 'y'}
_FALSE = {'0', 'false', 'no', 'n'}


class RowError(Exception):
    pass


def parse_file(django_file, filename):
    """Return list of dict rows from an uploaded .xlsx or .csv file."""
    if filename.lower().endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(django_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip().lower() if h is not None else '' for h in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for r in rows_iter:
            if all(v is None or str(v).strip() == '' for v in r):
                continue
            rows.append(dict(zip(headers, r)))
        wb.close()
        return rows
    # CSV
    import csv
    import io
    text = django_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    reader.fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
    return [row for row in reader if any((v or '').strip() for v in row.values())]


def _cell(row, key):
    """Normalized cell value: stripped string, or None when blank/missing."""
    v = row.get(key)
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None


def _dec(value, field):
    try:
        d = Decimal(value)
    except (InvalidOperation, ValueError):
        raise RowError(f'{field}: "{value}" is not a number')
    if d < 0:
        raise RowError(f'{field}: must be >= 0')
    return d


def _int(value, field):
    try:
        i = int(float(value))
    except (ValueError, TypeError):
        raise RowError(f'{field}: "{value}" is not an integer')
    if i < 0:
        raise RowError(f'{field}: must be >= 0')
    return i


def _bool(value, field):
    s = str(value).strip().lower()
    if s in _TRUE:
        return True
    if s in _FALSE:
        return False
    raise RowError(f'{field}: "{value}" is not a valid boolean (use 1/0/true/false)')


def _build_updates(row):
    """Validate a row and return dict of model-field updates (blank = absent)."""
    updates = {}
    for f in ('name_ar', 'name_en', 'description_ar', 'description_en'):
        v = _cell(row, f)
        if v is not None:
            updates[f] = v
    for f in ('original_price', 'discount_price', 'cost_price'):
        v = _cell(row, f)
        if v is not None:
            updates[f] = _dec(v, f)
    for f in ('quantity_in_stock', 'low_stock_threshold'):
        v = _cell(row, f)
        if v is not None:
            updates[f] = _int(v, f)
    v = _cell(row, 'sell_unit')
    if v is not None:
        unit = v.lower()
        valid = [c[0] for c in Product.SellUnit.choices]
        if unit not in valid:
            raise RowError(f'sell_unit: "{v}" is not one of {", ".join(valid)}')
        updates['sell_unit'] = unit
    for f in ('is_weight_based', 'is_available'):
        v = _cell(row, f)
        if v is not None:
            updates[f] = _bool(v, f)
    v = _cell(row, 'image_url')
    if v is not None:
        if not v.lower().startswith(('http://', 'https://')):
            raise RowError(f'image_url: "{v}" is not a valid URL')
        updates['image_url_s3'] = v
    return updates


def _match_categories(names_cell, category_index):
    """Return ([Category], [unmatched names]) for a comma-separated cell."""
    matched, missing = [], []
    for name in names_cell.split(','):
        name = name.strip()
        if not name:
            continue
        cat = category_index.get(name.lower())
        if cat is not None:
            matched.append(cat)
        else:
            missing.append(name)
    return matched, missing


def run_import(rows, user, store_id, dry_run=False):
    """Validate all rows and (unless dry_run) upsert products by barcode.

    Writes are batched: every row is validated first (no DB writes), then all
    changes are flushed with bulk_update / bulk_create and a single pair of
    queries for the category M2M, inside one transaction. This keeps a
    full-catalog import (10k+ rows) well under the request/worker timeout that
    a per-row save() loop would blow past.
    """
    result = {'total_rows': len(rows), 'created': 0, 'updated': 0, 'errors': [], 'warnings': []}
    if len(rows) > MAX_ROWS:
        result['errors'].append({'row': 0, 'barcode': '', 'reason': f'File has {len(rows)} rows; maximum is {MAX_ROWS}'})
        return result

    # Products visible to this admin (store-scoped for store admins).
    product_qs = Product.objects.all()
    user_store_id = getattr(user, 'store_id', None)
    if user_store_id is not None:
        product_qs = product_qs.filter(store_id=user_store_id)
        store_id = user_store_id
    if store_id is None:
        from apps.stores.models import Store
        first = Store.objects.order_by('id').first()
        store_id = first.id if first else None

    cat_qs = Category.objects.all()
    if store_id is not None:
        cat_qs = cat_qs.filter(store_id=store_id)
    category_index = {}
    for c in cat_qs:
        category_index[c.name_en.strip().lower()] = c
        category_index[c.name_ar.strip().lower()] = c

    barcodes = [b for b in (_cell(r, 'barcode') for r in rows) if b]
    existing = {p.barcode: p for p in product_qs.filter(barcode__in=barcodes)}
    # Barcode is globally unique; store admins may not touch another store's.
    other_store_barcodes = set()
    if user_store_id is not None:
        other_store_barcodes = set(
            Product.objects.filter(barcode__in=barcodes)
            .exclude(store_id=user_store_id)
            .values_list('barcode', flat=True)
        )

    to_update = []                 # existing Product instances, mutated
    to_create = []                 # new Product instances
    cat_assignments = []           # (Product, [Category]) — set() semantics, applied in bulk

    seen = set()
    for idx, row in enumerate(rows):
        rownum = idx + 2  # 1-based + header row
        barcode = _cell(row, 'barcode')
        try:
            if not barcode:
                raise RowError('barcode is required')
            if barcode in seen:
                raise RowError('duplicate barcode in file')
            seen.add(barcode)

            updates = _build_updates(row)
            categories, missing = None, []
            cats_cell = _cell(row, 'categories')
            if cats_cell is not None:
                categories, missing = _match_categories(cats_cell, category_index)

            product = existing.get(barcode)
            if product is None:
                if user_store_id is not None and barcode in other_store_barcodes:
                    raise RowError('barcode belongs to a product in another store')
                for f in REQUIRED_FOR_CREATE:
                    if f not in updates:
                        raise RowError(f'{f} is required for new products')
                if store_id is None:
                    raise RowError('no store configured — cannot create products')
                if not dry_run:
                    product = Product(store_id=store_id, barcode=barcode, name_en=updates.get('name_en', ''))
                    for f, v in updates.items():
                        setattr(product, f, v)
                    to_create.append(product)
                    if categories:
                        cat_assignments.append((product, categories))
                result['created'] += 1
            else:
                if not dry_run:
                    if 'image_url_s3' in updates:
                        # New URL image replaces any uploaded file image.
                        product.main_image = None
                    for f, v in updates.items():
                        setattr(product, f, v)
                    to_update.append(product)
                    if categories is not None:
                        cat_assignments.append((product, categories))
                result['updated'] += 1

            for name in missing:
                result['warnings'].append({
                    'row': rownum, 'barcode': barcode,
                    'reason': f'category "{name}" not found — skipped',
                })
        except RowError as e:
            result['errors'].append({'row': rownum, 'barcode': barcode or '', 'reason': str(e)})
        except Exception as e:  # unexpected validation failure on this row
            result['errors'].append({'row': rownum, 'barcode': barcode or '', 'reason': f'unexpected error: {e}'})

    if not dry_run and (to_update or to_create):
        try:
            with transaction.atomic():
                if to_update:
                    Product.objects.bulk_update(to_update, UPDATE_FIELDS, batch_size=500)
                if to_create:
                    Product.objects.bulk_create(to_create, batch_size=500)
                if cat_assignments:
                    Through = Product.categories.through
                    pids = [p.pk for p, _ in cat_assignments]
                    Through.objects.filter(product_id__in=pids).delete()
                    through_rows = [
                        Through(product_id=p.pk, category_id=c.pk)
                        for p, cats in cat_assignments for c in cats
                    ]
                    if through_rows:
                        Through.objects.bulk_create(through_rows, batch_size=1000, ignore_conflicts=True)
        except Exception as e:  # whole batch rolled back — nothing was written
            result['errors'].append({'row': 0, 'barcode': '', 'reason': f'bulk write failed, no changes saved: {e}'})
            result['created'] = 0
            result['updated'] = 0
    return result


def build_template_workbook(products=None):
    """Return an openpyxl Workbook: header row + optional existing products."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 4)

    if products:
        for r, p in enumerate(products, start=2):
            cats = ', '.join(c.name_en or c.name_ar for c in p.categories.all())
            image = p.image_url_s3 or (p.main_image.url if p.main_image else '')
            values = [
                p.barcode or '', p.name_ar, p.name_en, p.description_ar, p.description_en,
                cats, p.original_price, p.discount_price, p.cost_price,
                p.quantity_in_stock, p.low_stock_threshold, p.sell_unit,
                int(p.is_weight_based), int(p.is_available), image,
            ]
            for col, v in enumerate(values, start=1):
                ws.cell(row=r, column=col, value=v)
    return wb
