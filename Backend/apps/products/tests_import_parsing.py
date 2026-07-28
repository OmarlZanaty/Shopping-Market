# -*- coding: utf-8 -*-
"""
Tests for the product-import spreadsheet parser.

Deliberately Django-free (import_parsing has no Django imports) so it runs
anywhere openpyxl is installed:

    python apps/products/tests_import_parsing.py
"""
import io
import os
import re
import sys
import unittest
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl  # noqa: E402

from import_parsing import FORMULA_NO_VALUE, ROW_NUMBER_KEY, parse_file  # noqa: E402

HDR = ['barcode', 'name_ar', 'name_en', 'categories', 'original_price']


def xlsx(sheets):
    """Build an in-memory .xlsx. `sheets` is [(title, [rows])]."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ParseFileTests(unittest.TestCase):

    def test_formula_without_cached_value_is_flagged_not_silently_blank(self):
        """The reported bug: English names filled by a formula vanished."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HDR)
        ws['A2'], ws['B2'], ws['D2'], ws['E2'] = '111', 'اسم', 'Snacks', 10
        ws['C2'] = '=CONCATENATE("Pepsi"," 1L")'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        row = parse_file(buf, 'f.xlsx').rows[0]
        self.assertIs(row['name_en'], FORMULA_NO_VALUE)
        self.assertEqual(row['barcode'], '111')

    def test_data_on_a_non_active_sheet_is_found(self):
        f = xlsx([('Notes', [['read me']]),
                  ('Products', [HDR, ['222', 'اسم', 'Name', 'Snacks', 5]])])
        res = parse_file(f, 'f.xlsx')
        self.assertEqual(res.sheet, 'Products')
        self.assertEqual(len(res.rows), 1)

    def test_sheet_with_most_rows_wins(self):
        f = xlsx([('Sample', [HDR, ['1', 'a', 'A', 'c', 1]]),
                  ('Full', [HDR] + [[str(i), 'a', 'A', 'c', 1] for i in range(10)])])
        self.assertEqual(parse_file(f, 'f.xlsx').sheet, 'Full')

    def test_header_below_title_rows(self):
        f = xlsx([('S', [['بيانات الاصناف'], [], HDR, ['333', 'اسم', 'N', 'c', 3]])])
        res = parse_file(f, 'f.xlsx')
        self.assertEqual(res.header_row, 3)
        self.assertEqual(res.rows[0]['barcode'], '333')

    def test_client_column_names_are_aliased(self):
        f = xlsx([('S', [['bracode1', 'Item Name AR', 'English Name', 'ItemGroup', 'Sale_Price1'],
                         ['444', 'اسم', 'Chips', 'شيبس و سناكس', 12.5]])])
        res = parse_file(f, 'f.xlsx')
        row = res.rows[0]
        self.assertEqual(row['barcode'], '444')
        self.assertEqual(row['name_en'], 'Chips')
        self.assertEqual(row['categories'], 'شيبس و سناكس')
        self.assertEqual(row['original_price'], 12.5)
        # 'Item Name AR' is not an alias we accept — it must be reported.
        self.assertIn('Item Name AR', res.unknown_columns)

    def test_unknown_and_missing_columns_are_reported(self):
        f = xlsx([('S', [['barcode', 'name_en', 'Colour'], ['555', 'N', 'red']])])
        res = parse_file(f, 'f.xlsx')
        self.assertEqual(res.unknown_columns, ['Colour'])
        self.assertIn('name_ar', res.missing_columns)
        self.assertNotIn('name_en', res.missing_columns)

    def test_duplicate_column_keeps_first_non_blank(self):
        f = xlsx([('S', [['barcode', 'name_en', 'english name'],
                         ['666', 'Real Name', None],
                         ['777', None, 'Fallback']])])
        res = parse_file(f, 'f.xlsx')
        self.assertEqual(res.rows[0]['name_en'], 'Real Name')
        self.assertEqual(res.rows[1]['name_en'], 'Fallback')
        self.assertEqual(res.duplicate_columns, ['name_en'])

    def test_short_rows_are_padded_not_truncated(self):
        """Files declaring a narrow <dimension> used to lose trailing columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HDR)
        ws.append(['888', 'اسم', 'Name', 'Snacks', 9])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        narrowed = io.BytesIO()
        zin = zipfile.ZipFile(buf)
        zout = zipfile.ZipFile(narrowed, 'w', zipfile.ZIP_DEFLATED)
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                data = re.sub(rb'<dimension[^/]*/>', b'<dimension ref="A1:C2"/>', data)
            zout.writestr(item, data)
        zout.close()
        zin.close()
        narrowed.seek(0)

        row = parse_file(narrowed, 'f.xlsx').rows[0]
        self.assertEqual(row['barcode'], '888')
        self.assertIsNone(row.get('original_price'))  # absent, not a wrong value

    def test_blank_rows_are_skipped_but_row_numbers_stay_true(self):
        f = xlsx([('S', [HDR, ['1', 'a', 'A', 'c', 1], [None] * 5, ['2', 'b', 'B', 'c', 2]])])
        rows = parse_file(f, 'f.xlsx').rows
        self.assertEqual(len(rows), 2)
        self.assertEqual([r[ROW_NUMBER_KEY] for r in rows], [2, 4])

    def test_file_without_a_barcode_column_reports_nothing_found(self):
        f = xlsx([('S', [['name_ar', 'price'], ['اسم', 3]])])
        res = parse_file(f, 'f.xlsx')
        self.assertEqual(res.rows, [])
        self.assertEqual(res.sheet, '')
        self.assertEqual(res.sheets_scanned, ['S'])

    def test_csv_with_bom_and_aliases(self):
        data = '﻿Barcode,English Name,Sale_Price1\n999,Water,4\n'.encode('utf-8')
        res = parse_file(io.BytesIO(data), 'f.csv')
        self.assertEqual(res.rows[0]['barcode'], '999')
        self.assertEqual(res.rows[0]['name_en'], 'Water')
        self.assertEqual(res.rows[0]['original_price'], '4')

    def test_csv_with_title_row_above_header(self):
        data = 'Product export\n\nbarcode,name_en\n123,Milk\n'.encode('utf-8')
        res = parse_file(io.BytesIO(data), 'f.csv')
        self.assertEqual(res.header_row, 3)
        self.assertEqual(res.rows[0]['name_en'], 'Milk')


if __name__ == '__main__':
    unittest.main(verbosity=2)
