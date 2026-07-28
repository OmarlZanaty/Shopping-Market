"""
Spreadsheet parsing for the product importer — deliberately Django-free so it
can be unit-tested without a settings module or a database.

What this layer is responsible for (and why each piece exists):

* **Sheet selection.** `wb.active` is whatever sheet was selected when the file
  was last saved. Client workbooks routinely carry a cover/notes sheet, so we
  scan every sheet and pick the one that actually looks like product data.
* **Header row detection.** The header is not always row 1 — files exported
  from a POS often carry a title and a blank row above it. We scan the first
  `HEADER_SCAN_ROWS` rows of each sheet for a row containing a barcode column.
* **Header aliasing.** `bracode1`, `Sale_Price1`, `English Name` … all mean
  something. Anything we still can't place is reported as an unknown column
  instead of being silently dropped.
* **Formula cells with no cached value.** openpyxl reads values, not formulas.
  A file whose English names were filled by a formula and was last saved by
  something other than Excel has no cached result, so the cell reads as blank —
  which the importer used to treat as "leave unchanged", updating the row and
  reporting success while the column never changed. Those cells are tagged with
  `FORMULA_NO_VALUE` so the row fails loudly.
* **Short rows.** Some writers emit a narrow `<dimension>`; zipping headers
  against a shorter row silently truncated the trailing columns. Rows are
  padded to the header width.
"""
import re

MAX_ROWS = 50000
HEADER_SCAN_ROWS = 15

# Reserved key carrying each row's real 1-based sheet row number. It can never
# collide with a parsed column: column keys come from ALIASES.
ROW_NUMBER_KEY = '__sheet_row__'

# Canonical import columns, in template order.
COLUMNS = [
    'barcode', 'name_ar', 'name_en', 'description_ar', 'description_en',
    'categories', 'original_price', 'discount_price', 'cost_price',
    'quantity_in_stock', 'low_stock_threshold', 'sell_unit',
    'is_weight_based', 'is_available', 'image_url',
]


class _FormulaNoValue:
    """Marker for a formula cell whose cached result is missing."""

    __slots__ = ()

    def __repr__(self):  # pragma: no cover - debugging aid
        return '<FORMULA_NO_VALUE>'

    def __bool__(self):
        return False


FORMULA_NO_VALUE = _FormulaNoValue()


# Alias -> canonical column. Keys are already normalized (see _norm).
# Only unambiguous aliases belong here: a header like `item_name` could be
# either language, so it is left unknown and surfaced in the preview rather
# than guessed at.
ALIASES = {}


def _alias(canonical, *names):
    for n in names:
        ALIASES[_norm(n)] = canonical


def _norm(header):
    """Normalize a header cell to a comparison key: lowercase, _-joined."""
    if header is None:
        return ''
    s = str(header).replace(' ', ' ').replace('‏', '').replace('‎', '')
    s = s.strip().lower()
    s = re.sub(r'[\s\-\./\\]+', '_', s)
    return s.strip('_')


_alias('barcode', 'barcode', 'bar_code', 'barcode1', 'bracode1', 'bracode',
       'item_barcode', 'ean', 'ean13', 'upc', 'sku', 'باركود', 'الباركود')
_alias('name_ar', 'name_ar', 'namear', 'arabic_name', 'name_arabic', 'ar_name',
       'الاسم_العربي', 'الاسم_بالعربية', 'اسم_الصنف', 'اسم_المنتج')
_alias('name_en', 'name_en', 'nameen', 'english_name', 'name_english', 'en_name',
       'english', 'الاسم_الانجليزي', 'الاسم_بالانجليزية')
_alias('description_ar', 'description_ar', 'arabic_description', 'desc_ar', 'الوصف_العربي')
_alias('description_en', 'description_en', 'english_description', 'desc_en', 'الوصف_الانجليزي')
_alias('categories', 'categories', 'category', 'item_group', 'itemgroup', 'group',
       'department', 'section', 'القسم', 'الاقسام', 'المجموعة', 'التصنيف')
_alias('original_price', 'original_price', 'price', 'sale_price', 'sale_price1',
       'saleprice1', 'selling_price', 'unit_price', 'السعر', 'سعر_البيع')
_alias('discount_price', 'discount_price', 'discounted_price', 'offer_price',
       'special_price', 'سعر_الخصم', 'السعر_بعد_الخصم')
_alias('cost_price', 'cost_price', 'cost', 'purchase_price', 'buy_price', 'سعر_التكلفة')
_alias('quantity_in_stock', 'quantity_in_stock', 'quantity', 'qty', 'stock',
       'stock_qty', 'الكمية', 'المخزون')
_alias('low_stock_threshold', 'low_stock_threshold', 'low_stock', 'min_stock',
       'reorder_level', 'حد_الطلب')
_alias('sell_unit', 'sell_unit', 'unit', 'uom', 'الوحدة', 'وحدة_البيع')
_alias('is_weight_based', 'is_weight_based', 'weight_based', 'by_weight', 'بالوزن')
_alias('is_available', 'is_available', 'available', 'is_active', 'active', 'متاح')
_alias('image_url', 'image_url', 'image', 'image_link', 'img', 'img_url',
       'picture', 'photo', 'الصورة', 'رابط_الصورة')


class ParseResult:
    """Rows plus everything the preview needs to explain what was ignored."""

    def __init__(self, rows, sheet='', header_row=0, unknown_columns=None,
                 missing_columns=None, duplicate_columns=None, sheets_scanned=None):
        self.rows = rows
        self.sheet = sheet
        self.header_row = header_row
        self.unknown_columns = unknown_columns or []
        self.missing_columns = missing_columns or []
        self.duplicate_columns = duplicate_columns or []
        self.sheets_scanned = sheets_scanned or []

    def __len__(self):
        return len(self.rows)

    def __iter__(self):
        return iter(self.rows)

    def as_meta(self):
        return {
            'sheet': self.sheet,
            'header_row': self.header_row,
            'sheets_scanned': self.sheets_scanned,
            'unknown_columns': self.unknown_columns,
            'missing_columns': self.missing_columns,
            'duplicate_columns': self.duplicate_columns,
        }


def _is_blank(value):
    """A missing formula result is not blank — it must not be skipped."""
    if isinstance(value, _FormulaNoValue):
        return False
    return value is None or str(value).strip() == ''


def _blank_row(values):
    return all(v is None or str(v).strip() == '' for v in values)


def _find_header(grid):
    """Index of the first row that looks like a header, or None.

    A header row is one whose cells map to at least two known columns and
    include `barcode` — the upsert key, so a file without it is unusable
    anyway.
    """
    for idx, row in enumerate(grid[:HEADER_SCAN_ROWS]):
        mapped = {ALIASES.get(_norm(c)) for c in row}
        mapped.discard(None)
        if 'barcode' in mapped and len(mapped) >= 2:
            return idx
    return None


def _map_headers(header_cells):
    """(keys per column, unknown labels, duplicated canonical names).

    `keys` is per-column: the canonical name when recognized, otherwise ''
    so the column is dropped from the row dict but reported to the user.
    """
    keys, unknown, seen, duplicates = [], [], set(), []
    for cell in header_cells:
        raw = '' if cell is None else str(cell).strip()
        canonical = ALIASES.get(_norm(cell))
        if canonical is None:
            keys.append('')
            if raw:
                unknown.append(raw)
            continue
        if canonical in seen:
            duplicates.append(canonical)
        seen.add(canonical)
        keys.append(canonical)
    return keys, unknown, duplicates


def _rows_from_grid(grid, header_idx):
    """Build row dicts, padding short rows and keeping the first value when a
    canonical column appears more than once."""
    keys, unknown, duplicates = _map_headers(grid[header_idx])
    width = len(keys)
    rows = []
    for sheet_row, values in enumerate(grid[header_idx + 1:], start=header_idx + 2):
        if _blank_row(values):
            continue
        values = list(values) + [None] * (width - len(values))
        # Real sheet row number, so an error points at the row the client sees
        # even when blank rows were skipped along the way.
        row = {ROW_NUMBER_KEY: sheet_row}
        for key, value in zip(keys, values):
            if not key:
                continue
            if key in row and (_is_blank(value) or not _is_blank(row[key])):
                continue  # duplicated column: first non-blank value wins
            row[key] = value
        rows.append(row)
    missing = [c for c in COLUMNS if c not in set(k for k in keys if k)]
    return rows, unknown, duplicates, missing


def _sheet_grid(ws_values, ws_formulas):
    """Materialize a sheet as a list of row tuples, tagging formula cells whose
    cached value is missing with FORMULA_NO_VALUE."""
    formula_rows = ws_formulas.iter_rows(values_only=True) if ws_formulas is not None else None
    grid = []
    for values in ws_values.iter_rows(values_only=True):
        formulas = ()
        if formula_rows is not None:
            try:
                formulas = next(formula_rows)
            except StopIteration:
                formulas = ()
        row = []
        for i, v in enumerate(values):
            if v is None and i < len(formulas):
                f = formulas[i]
                if isinstance(f, str) and f.startswith('='):
                    v = FORMULA_NO_VALUE
            row.append(v)
        grid.append(tuple(row))
        if len(grid) > MAX_ROWS + HEADER_SCAN_ROWS:
            break
    return grid


def parse_xlsx(django_file):
    """Parse an .xlsx across all sheets; return the best ParseResult."""
    import io

    import openpyxl

    # Both workbooks are read lazily, so they cannot share one file handle —
    # two ZipFile readers seeking the same descriptor corrupt each other. Buffer
    # the upload once and hand each reader its own view of the bytes.
    django_file.seek(0)
    payload = django_file.read()

    wb_values = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        wb_formulas = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    except Exception:  # unreadable second pass; formula detection unavailable
        wb_formulas = None

    best = None
    scanned = []
    try:
        for name in wb_values.sheetnames:
            ws_values = wb_values[name]
            ws_formulas = wb_formulas[name] if wb_formulas is not None and name in wb_formulas.sheetnames else None
            grid = _sheet_grid(ws_values, ws_formulas)
            header_idx = _find_header(grid)
            scanned.append(name)
            if header_idx is None:
                continue
            rows, unknown, duplicates, missing = _rows_from_grid(grid, header_idx)
            candidate = ParseResult(
                rows, sheet=name, header_row=header_idx + 1,
                unknown_columns=unknown, missing_columns=missing,
                duplicate_columns=duplicates, sheets_scanned=scanned,
            )
            # More data rows wins: a stray sheet with a copied header row
            # should never beat the real catalogue.
            if best is None or len(candidate.rows) > len(best.rows):
                best = candidate
    finally:
        wb_values.close()
        if wb_formulas is not None:
            wb_formulas.close()

    if best is None:
        return ParseResult([], sheets_scanned=scanned)
    best.sheets_scanned = scanned
    return best


def parse_csv(django_file):
    import csv
    import io

    django_file.seek(0)
    raw = django_file.read()
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8-sig', errors='replace')
    grid = [tuple(r) for r in csv.reader(io.StringIO(raw))]
    header_idx = _find_header(grid)
    if header_idx is None:
        return ParseResult([], sheet='CSV')
    rows, unknown, duplicates, missing = _rows_from_grid(grid, header_idx)
    return ParseResult(rows, sheet='CSV', header_row=header_idx + 1,
                       unknown_columns=unknown, missing_columns=missing,
                       duplicate_columns=duplicates)


def parse_file(django_file, filename):
    """Return a ParseResult for an uploaded .xlsx or .csv file."""
    if filename.lower().endswith('.xlsx'):
        return parse_xlsx(django_file)
    return parse_csv(django_file)
