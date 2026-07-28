# -*- coding: utf-8 -*-
"""
End-to-end tests for the product import: run_import against the database and
the admin upload endpoint.

    DJANGO_SETTINGS_MODULE=config.settings_test python manage.py test apps.products
"""
import io

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIClient

from apps.stores.models import Store

from . import importer
from .models import Category, Product

HDR = ['barcode', 'name_ar', 'name_en', 'categories', 'original_price']


def make_xlsx(sheets):
    """Build an in-memory .xlsx. `sheets` is [(title, [rows])]."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ImportTestBase(TestCase):

    def setUp(self):
        self.store = Store.objects.create(name_ar='متجر', name_en='Store', type='supermarket')
        self.category = Category.objects.create(
            store=self.store, name_ar='شيبس و سناكس', name_en='Snacks')
        self.admin = get_user_model().objects.create_user(
            phone='+201000000001', full_name='Admin', role='admin',
            is_staff=True, is_superuser=True, is_active=True,
        )

    def upload(self, buf, filename='import.xlsx', dry_run=False):
        client = APIClient()
        client.force_authenticate(user=self.admin)
        return client.post(
            '/api/v1/products/admin/products/import/',
            {'file': SimpleUploadedFile(filename, buf.getvalue()),
             'dry_run': 'true' if dry_run else 'false',
             'store_id': str(self.store.id)},
            format='multipart',
        )


class RunImportTests(ImportTestBase):

    def test_creates_and_updates_by_barcode(self):
        f = make_xlsx([('S', [HDR, ['1001', 'شيبس', 'Chips', 'Snacks', 15]])])
        res = importer.run_import(importer.parse_file(f, 'a.xlsx'), self.admin, self.store.id)
        self.assertEqual((res['created'], res['updated'], res['errors']), (1, 0, []))

        p = Product.objects.get(barcode='1001')
        self.assertEqual(p.name_en, 'Chips')
        self.assertEqual(str(p.original_price), '15.00')
        self.assertEqual(list(p.categories.values_list('id', flat=True)), [self.category.id])

        f = make_xlsx([('S', [HDR, ['1001', 'شيبس', 'Chips Large', 'Snacks', 18]])])
        res = importer.run_import(importer.parse_file(f, 'b.xlsx'), self.admin, self.store.id)
        self.assertEqual((res['created'], res['updated']), (0, 1))
        p.refresh_from_db()
        self.assertEqual(p.name_en, 'Chips Large')
        self.assertEqual(str(p.original_price), '18.00')

    def test_formula_name_en_fails_the_row_instead_of_silently_skipping(self):
        """The reported bug: the row updated, the English name did not."""
        Product.objects.create(store=self.store, barcode='1002', name_ar='قديم',
                               name_en='Old Name', original_price=5)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HDR)
        ws['A2'], ws['B2'], ws['D2'], ws['E2'] = '1002', 'جديد', 'Snacks', 7
        ws['C2'] = '=B2&" EN"'          # formula, no cached result
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = importer.run_import(importer.parse_file(buf, 'c.xlsx'), self.admin, self.store.id)
        self.assertEqual(res['updated'], 0)
        self.assertEqual(len(res['errors']), 1)
        self.assertIn('name_en', res['errors'][0]['reason'])
        self.assertIn('formula', res['errors'][0]['reason'])
        self.assertEqual(res['errors'][0]['row'], 2)

        p = Product.objects.get(barcode='1002')
        self.assertEqual(p.name_en, 'Old Name')   # untouched, and we said so
        self.assertEqual(p.name_ar, 'قديم')       # whole row rejected, not half

    def _client_workbook(self, name_ar_header):
        """The shape of the client's real workbook: a categories sheet first,
        a title row above the headers, and POS-style column names."""
        return make_xlsx([
            ('الاقسام', [['الاقسام'], [], ['م', 'اسم القسم'], [1, 'شيبس و سناكس']]),
            ('الاصناف', [['بيانات الاصناف'], [],
                         ['bracode1', name_ar_header, 'English Name', 'ItemGroup', 'Sale_Price1'],
                         ['1003', 'شيبس', 'Chips', 'شيبس و سناكس', 12.5]]),
        ])

    def test_client_workbook_shape_imports(self):
        f = self._client_workbook('name_ar')
        parsed = importer.parse_file(f, 'd.xlsx')
        self.assertEqual((parsed.sheet, parsed.header_row), ('الاصناف', 3))

        res = importer.run_import(parsed, self.admin, self.store.id)
        self.assertEqual((res['created'], res['errors']), (1, []))
        p = Product.objects.get(barcode='1003')
        self.assertEqual((p.name_ar, p.name_en), ('شيبس', 'Chips'))
        self.assertEqual(str(p.original_price), '12.50')
        self.assertEqual(p.categories.first().name_ar, 'شيبس و سناكس')

    def test_ambiguous_name_column_is_refused_not_guessed(self):
        """`Item_Name` could be either language. Writing it into name_ar on a
        hunch is worse than failing, so the row fails and the column is named."""
        f = self._client_workbook('Item_Name')
        parsed = importer.parse_file(f, 'd.xlsx')
        self.assertIn('Item_Name', parsed.unknown_columns)

        res = importer.run_import(parsed, self.admin, self.store.id)
        self.assertEqual(res['created'], 0)
        self.assertIn('name_ar is required', res['errors'][0]['reason'])
        self.assertFalse(Product.objects.filter(barcode='1003').exists())

    def test_blank_cell_still_keeps_the_existing_value(self):
        Product.objects.create(store=self.store, barcode='1004', name_ar='اسم',
                               name_en='Keep Me', original_price=5)
        f = make_xlsx([('S', [HDR, ['1004', 'اسم', None, 'Snacks', 9]])])
        res = importer.run_import(importer.parse_file(f, 'e.xlsx'), self.admin, self.store.id)
        self.assertEqual(res['updated'], 1)
        p = Product.objects.get(barcode='1004')
        self.assertEqual(p.name_en, 'Keep Me')
        self.assertEqual(str(p.original_price), '9.00')

    def test_dry_run_writes_nothing(self):
        f = make_xlsx([('S', [HDR, ['1005', 'اسم', 'Name', 'Snacks', 4]])])
        res = importer.run_import(importer.parse_file(f, 'g.xlsx'), self.admin,
                                  self.store.id, dry_run=True)
        self.assertEqual(res['created'], 1)
        self.assertFalse(Product.objects.filter(barcode='1005').exists())

    def test_error_rows_report_the_real_sheet_row(self):
        f = make_xlsx([('S', [['title'], [], HDR,
                              ['1006', 'اسم', 'A', 'Snacks', 3],
                              [None] * 5,
                              [None, 'اسم', 'B', 'Snacks', 3]])])   # missing barcode
        res = importer.run_import(importer.parse_file(f, 'h.xlsx'), self.admin, self.store.id)
        self.assertEqual(res['created'], 1)
        self.assertEqual(len(res['errors']), 1)
        self.assertEqual(res['errors'][0]['row'], 6)   # the row the client sees

    def test_large_file_imports_in_one_batch(self):
        rows = [[str(90000 + i), f'اسم {i}', f'Name {i}', 'Snacks', 5] for i in range(2000)]
        f = make_xlsx([('S', [HDR] + rows)])
        res = importer.run_import(importer.parse_file(f, 'i.xlsx'), self.admin, self.store.id)
        self.assertEqual((res['created'], res['errors']), (2000, []))
        self.assertEqual(Product.objects.filter(store=self.store).count(), 2000)


class ImportEndpointTests(ImportTestBase):

    def test_upload_reports_sheet_and_ignored_columns(self):
        f = make_xlsx([('Notes', [['cover page']]),
                       ('Data', [['barcode', 'name_ar', 'English Name', 'Colour', 'price'],
                                 ['2001', 'اسم', 'Name', 'red', 6]])])
        resp = self.upload(f, dry_run=True)
        self.assertEqual(resp.status_code, 200, resp.content)
        data = resp.json()['data']

        self.assertEqual(data['created'], 1)
        self.assertEqual(data['file']['sheet'], 'Data')
        self.assertEqual(data['file']['header_row'], 1)
        self.assertEqual(data['file']['unknown_columns'], ['Colour'])
        self.assertTrue(any('Colour' in w['reason'] for w in data['warnings']))
        self.assertFalse(Product.objects.filter(barcode='2001').exists())  # dry run

    def test_upload_without_a_barcode_column_is_rejected_with_a_clear_message(self):
        f = make_xlsx([('S', [['name_ar', 'price'], ['اسم', 3]])])
        resp = self.upload(f)
        self.assertEqual(resp.status_code, 400)
        self.assertIn('barcode', str(resp.content, 'utf-8'))

    def test_real_upload_writes_and_records_a_job(self):
        from .models import ProductImportJob
        f = make_xlsx([('S', [HDR, ['2002', 'اسم', 'Name', 'Snacks', 8]])])
        resp = self.upload(f)
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertTrue(Product.objects.filter(barcode='2002').exists())
        self.assertEqual(ProductImportJob.objects.count(), 1)
