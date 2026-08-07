# -*- coding: utf-8 -*-
"""
Tests for the reports the dashboard's Reports page calls.

Covers the report the owner asked for that did not exist (sales-summary) and
the export path, because a report that renders but cannot be exported is not
the feature that was requested.

    DJANGO_SETTINGS_MODULE=config.settings_test python manage.py test apps.analytics
"""
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from apps.branches.models import Branch
from apps.orders.models import Order
from apps.stores.models import Store

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class SalesSummaryReportTests(TestCase):

    def setUp(self):
        self.store = Store.objects.create(name_ar='متجر', name_en='Store', type='supermarket')
        self.branch = Branch.objects.create(
            store=self.store, name='Main', name_ar='الفرع الرئيسي', name_en='Main',
            address='...', latitude=27.9, longitude=34.3, phone='0100',
        )
        User = get_user_model()
        self.admin = User.objects.create_user(
            phone='+201000000001', full_name='Admin', role='admin',
            is_staff=True, is_superuser=True, is_active=True,
        )
        self.customer = User.objects.create_user(
            phone='+201000000002', full_name='محمد احمد', role='customer', is_active=True,
        )
        self.preparer = User.objects.create_user(
            phone='+201000000003', full_name='ابانوب', role='preparer', is_active=True,
        )

        now = timezone.now()
        self.order = Order.objects.create(
            store=self.store, branch=self.branch, customer=self.customer,
            preparer=self.preparer, status='delivered',
            delivery_address='بيلافيستا فيلا 2',
            total_amount=308.25, payment_method='cash', points_earned=10,
            accepted_at=now - timedelta(minutes=10),
            out_for_delivery_at=now,
            delivered_at=now,
        )

    def get(self, params=None):
        client = APIClient()
        client.force_authenticate(user=self.admin)
        return client.get('/api/v1/reports/sales-summary/', params or {})

    def test_row_carries_every_column_the_owner_asked_for(self):
        res = self.get()
        self.assertEqual(res.status_code, 200)
        rows = res.data['data']
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row['order_number'], self.order.order_number)
        self.assertEqual(row['amount'], 308.25)
        self.assertEqual(row['payment_method'], 'cash')
        self.assertEqual(row['preparer'], 'ابانوب')
        self.assertEqual(row['points_earned'], 10)
        self.assertEqual(row['items_count'], 0)
        # accepted_at → out_for_delivery_at, to one decimal
        self.assertAlmostEqual(row['prep_mins'], 10.0, places=0)

    def test_order_outside_the_range_is_excluded(self):
        old = (timezone.now() - timedelta(days=90)).date().isoformat()
        res = self.get({'from_date': old, 'to_date': old})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['data'], [])

    def test_unprepared_order_reports_a_blank_time_not_a_zero(self):
        self.order.accepted_at = None
        self.order.save(update_fields=['accepted_at'])
        row = self.get().data['data'][0]
        self.assertEqual(row['prep_mins'], '')

    def test_exports_as_xlsx(self):
        res = self.get({'export': 'xlsx'})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], XLSX_MIME)
        self.assertIn('sales_summary', res['Content-Disposition'])
        self.assertTrue(res.content.startswith(b'PK'))  # a real zip/xlsx

    def test_exports_as_pdf(self):
        res = self.get({'export': 'pdf'})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], 'application/pdf')
        self.assertTrue(res.content.startswith(b'%PDF'))


class ExportAllTests(TestCase):
    """One workbook, a sheet per report — the owner's "كل التقارير في ملف واحد"."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(
            phone='+201000000005', full_name='Admin', role='admin',
            is_staff=True, is_superuser=True, is_active=True,
        )
        self.client_ = APIClient()
        self.client_.force_authenticate(user=self.admin)

    def test_workbook_has_a_sheet_for_every_report(self):
        import io
        import openpyxl
        from apps.analytics.reports.views import ALL_REPORTS

        res = self.client_.get('/api/v1/reports/export-all/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], XLSX_MIME)

        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), len(ALL_REPORTS))
        self.assertIn('sales_summary', wb.sheetnames)
        self.assertIn('out_of_stock', wb.sheetnames)
        # Sheets are headed even when the range holds no rows.
        self.assertEqual(wb['sales_summary'].cell(row=1, column=1).value, 'Order #')

    def test_sheet_names_stay_within_excels_limits(self):
        import io
        import openpyxl

        res = self.client_.get('/api/v1/reports/export-all/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), len(set(wb.sheetnames)))
        for name in wb.sheetnames:
            self.assertLessEqual(len(name), 31)
            self.assertFalse(set(name) & set('[]:*?/\\'))


class ExistingReportsStillReachableTests(TestCase):
    """The other two sheets the owner drew already existed — keep them routed."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(
            phone='+201000000004', full_name='Admin', role='admin',
            is_staff=True, is_superuser=True, is_active=True,
        )

    def test_out_of_stock_and_sales_detail_answer(self):
        client = APIClient()
        client.force_authenticate(user=self.admin)
        for slug in ('out-of-stock', 'sales'):
            with self.subTest(slug=slug):
                res = client.get(f'/api/v1/reports/{slug}/')
                self.assertEqual(res.status_code, 200)
                self.assertIn('data', res.data)
