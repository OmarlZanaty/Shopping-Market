"""
Celery tasks for orders.

Includes:
- auto_close_order        — 2-hour auto-close
- approval_timeout_remind — 15-min approval timeout → push "call customer"
- broadcast_driver_location
- notify_stock_waitlist
- expire_discounts        — beat
- send_smart_notifications — beat
- process_product_import  — bulk Excel/CSV import
"""
import logging
from celery import shared_task
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task
def auto_close_order(order_id):
    """Auto-close order 2 hours after delivery mark if customer didn't confirm."""
    from .models import Order, SmartTimerAutoClose
    try:
        order = Order.objects.get(id=order_id)
        timer = SmartTimerAutoClose.objects.get(order=order, is_resolved=False)
        if order.status == Order.Status.OUT_FOR_DELIVERY:
            order.closed_by = Order.ClosedBy.DRIVER
            order.closed_by_driver_at = timezone.now()
            order.save(update_fields=['closed_by', 'closed_by_driver_at'])
            order.update_status(Order.Status.DELIVERED)
            order.award_points()
            timer.is_resolved = True
            timer.save()
            logger.info('order %s auto-closed after 2h', order.order_number)
    except Exception as e:
        logger.exception('auto_close error for %s: %s', order_id, e)


@shared_task
def approval_timeout_remind(adjustment_id):
    """
    Spec: 15-min timer. If customer didn't respond to a price/weight/substitute,
    push a "call customer" event to the agent's app so they can ring up.
    """
    from .models import OrderAdjustment
    from apps.notifications.utils import send_push_notification
    try:
        adj = OrderAdjustment.objects.select_related('order', 'preparer', 'driver').get(pk=adjustment_id)
    except OrderAdjustment.DoesNotExist:
        return
    if adj.customer_approval_status in ('approved', 'rejected'):
        return
    agent = adj.preparer or adj.driver
    if not agent:
        return
    send_push_notification(
        user=agent,
        title_ar='العميل لم يرد — اتصل الآن',
        title_en='Customer not responding — call now',
        body_ar=f'انتظر التأكيد لمدة 15 دقيقة على طلب {adj.order.order_number}',
        body_en=f'Awaiting customer reply on {adj.order.order_number} for 15 min',
        data={
            'type': 'call_customer',
            'order_id': str(adj.order.id),
            'order_number': adj.order.order_number,
            'adjustment_id': str(adj.id),
            'customer_phone': adj.order.delivery_phone,
        },
    )
    # Log the timeout firing on the adjustment trail
    OrderAdjustment.objects.create(
        order=adj.order, order_item=adj.order_item,
        preparer=adj.preparer, driver=adj.driver,
        action_type=OrderAdjustment.AdjustmentType.CALL_ATTEMPT,
        reason='15-min approval timeout',
    )
    logger.info('approval timeout fired for adjustment %s', adjustment_id)


@shared_task
def broadcast_driver_location(driver_id):
    """Broadcast driver location to relevant order rooms + admin dashboard."""
    from django.contrib.auth import get_user_model
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    User = get_user_model()
    try:
        driver = User.objects.get(id=driver_id)
        channel_layer = get_channel_layer()
        if not channel_layer:
            return
        from .models import Order
        active_orders = Order.objects.filter(
            driver=driver,
            status__in=['preparing', 'out_for_delivery'],
        )
        for order in active_orders:
            async_to_sync(channel_layer.group_send)(
                f'order_{order.order_number}',
                {
                    'type': 'driver_location',
                    'latitude': str(driver.current_latitude or 0),
                    'longitude': str(driver.current_longitude or 0),
                },
            )
        async_to_sync(channel_layer.group_send)(
            'admin_dashboard',
            {
                'type': 'driver_location_update',
                'driver_id': str(driver.id),
                'driver_name': driver.full_name,
                'latitude': str(driver.current_latitude or 0),
                'longitude': str(driver.current_longitude or 0),
            },
        )
    except Exception as e:
        logger.exception('broadcast_driver_location error: %s', e)


@shared_task
def notify_stock_waitlist(product_id):
    """Notify all waitlist subscribers when a product is back in stock."""
    from apps.products.models import StockWaitlist
    waitlist = StockWaitlist.objects.filter(
        product_id=product_id, notified_at__isnull=True
    ).select_related('user', 'product')
    for entry in waitlist:
        try:
            entry.notify()
        except Exception as e:
            logger.exception('waitlist notify error: %s', e)


@shared_task
def expire_discounts():
    """Beat-scheduled: deactivate expired discounts."""
    from apps.products.models import Product
    now = timezone.now()
    expired = Product.objects.filter(
        discount_end__lt=now, discount_price__isnull=False,
    )
    count = expired.count()
    expired.update(
        discount_price=None, discount_percentage=None,
        discount_start=None, discount_end=None,
    )
    if count:
        logger.info('expired %d product discounts', count)


@shared_task
def send_smart_notifications():
    """Daily: re-engagement push for customers inactive 7+ days."""
    from django.contrib.auth import get_user_model
    from apps.notifications.utils import send_push_notification
    from datetime import timedelta
    User = get_user_model()
    cutoff = timezone.now() - timedelta(days=7)
    inactive = User.objects.filter(
        role='customer', is_active=True,
        customer_orders__created_at__lt=cutoff,
    ).distinct()
    for user in inactive[:100]:
        send_push_notification(
            user=user,
            title_ar='اشتقنا إليك! 🛒',
            title_en='We miss you! 🛒',
            body_ar='عندنا عروض رائعة تنتظرك',
            body_en='Great deals waiting for you',
            data={'type': 'promotion'},
        )


@shared_task
def process_product_import(file_path, store_id):
    """Background bulk product import from CSV or XLSX."""
    import os
    from apps.products.models import Product
    success, errors = 0, 0
    rows = []
    try:
        if file_path.lower().endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            import csv
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
    except Exception as e:
        logger.exception('import parse error: %s', e)
        try:
            os.unlink(file_path)
        except OSError:
            pass
        return {'success': 0, 'errors': 1, 'message': str(e)}

    for row in rows:
        try:
            barcode = (row.get('barcode') or '').strip()
            name_ar = (row.get('name_ar') or '').strip()
            name_en = (row.get('name_en') or '').strip()
            if not name_ar or not row.get('original_price'):
                errors += 1
                continue
            payload = {
                'store_id': store_id,
                'barcode': barcode or None,
                'name_ar': name_ar,
                'name_en': name_en,
                'original_price': float(row['original_price']),
                'quantity_in_stock': int(row.get('quantity_in_stock') or 0),
                'sell_unit': row.get('unit_type') or 'piece',
                'is_available': str(row.get('is_available', '1')).strip() in ('1', 'true', 'True'),
            }
            if barcode:
                Product.objects.update_or_create(
                    barcode=barcode, defaults=payload,
                )
            else:
                Product.objects.create(**payload)
            success += 1
        except Exception as e:
            logger.warning('row import error: %s — %s', e, row)
            errors += 1

    try:
        os.unlink(file_path)
    except OSError:
        pass
    logger.info('product import done: %d ok, %d errors', success, errors)
    return {'success': success, 'errors': errors}


@shared_task
def update_order_streak(user_id):
    """Gamification — daily streak counter."""
    from django.contrib.auth import get_user_model
    from datetime import date, timedelta
    User = get_user_model()
    try:
        user = User.objects.get(id=user_id, role='customer')
        today = date.today()
        if user.last_order_date == today - timedelta(days=1):
            user.order_streak += 1
        elif user.last_order_date != today:
            user.order_streak = 1
        user.last_order_date = today
        user.save(update_fields=['order_streak', 'last_order_date'])
    except Exception as e:
        logger.exception('streak update error: %s', e)
