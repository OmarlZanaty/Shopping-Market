"""
Shared exporters: list-of-dicts → XLSX (openpyxl) or PDF (reportlab).

The PDF exporter wires `arabic-reshaper` + `python-bidi` when available so RTL
text renders correctly. If the deps are missing the PDF still generates but
without shaping — better than a 500 in production.
"""
import io
from datetime import datetime, date

from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_RTL = True
except ImportError:
    HAS_RTL = False


def _shape_ar(text):
    """Shape Arabic for proper RTL rendering. Safe fallback if libs missing."""
    if not text or not isinstance(text, str):
        return text
    if not HAS_RTL:
        return text
    try:
        return get_display(arabic_reshaper.reshape(text))
    except Exception:
        return text


def _fmt(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def export_xlsx(filename, columns, rows, sheet_name='Report'):
    """
    columns: [{ 'key': 'order_id', 'label': 'Order #', 'label_ar': 'رقم الطلب' }]
    rows: list of dicts keyed by column 'key'.
    Returns HttpResponse with xlsx content type.
    """
    if not HAS_XLSX:
        raise RuntimeError('openpyxl not installed. pip install openpyxl')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header row
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx, col in enumerate(columns, start=1):
        label = col.get('label') or col.get('label_ar') or col['key']
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(str(label)) + 6))

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_fmt(row.get(col['key'])))

    # Freeze the header
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = ''.join(c for c in filename if c.isalnum() or c in '-_') or 'report'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
    return response


def export_pdf(filename, columns, rows, title='Report', orientation='landscape'):
    if not HAS_PDF:
        raise RuntimeError('reportlab not installed. pip install reportlab')

    buf = io.BytesIO()
    pagesize = landscape(A4) if orientation == 'landscape' else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'title', parent=styles['Heading1'],
        fontSize=16, alignment=2,  # right
        textColor=colors.HexColor('#FF6B35'),
    )
    elements = [Paragraph(_shape_ar(title), title_style), Spacer(1, 6 * mm)]

    # Header row + data rows
    header = [_shape_ar(col.get('label') or col.get('label_ar') or col['key']) for col in columns]
    data = [header]
    for row in rows:
        data.append([_shape_ar(_fmt(row.get(col['key']))) for col in columns])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF6F0')]),
    ]))
    elements.append(table)
    doc.build(elements)

    buf.seek(0)
    safe_name = ''.join(c for c in filename if c.isalnum() or c in '-_') or 'report'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.pdf"'
    return response


def maybe_export(request, columns, rows, base_filename, title):
    """Dispatch on ?export=xlsx|pdf — returns None if no export requested."""
    fmt = (request.query_params.get('export') or '').lower()
    if fmt == 'xlsx':
        return export_xlsx(base_filename, columns, rows)
    if fmt == 'pdf':
        return export_pdf(base_filename, columns, rows, title=title)
    return None
